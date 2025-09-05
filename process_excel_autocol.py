import argparse, json, re
from pathlib import Path
from openpyxl import load_workbook

def load_jsonc(path: Path):
    t = path.read_text(encoding="utf-8")
    t = re.sub(r"//.*", "", t)
    t = re.sub(r"/\*.*?\*/", "", t, flags=re.DOTALL)
    return json.loads(t)

def to_float(x):
    try:
        if x is None: return None
        if isinstance(x, str):
            s = x.strip().replace(",", "")
            if s == "": return None
            return float(s)
        return float(x)
    except: return None

def safe_add(a,b): return (to_float(a) or 0.0) + (to_float(b) or 0.0)
def safe_sub(a,b): return (to_float(a) or 0.0) - (to_float(b) or 0.0)
def safe_div(n,d):
    n = to_float(n); d = to_float(d)
    if n is None or d in (None,0.0): return None
    return n/d

def normalize(s): return re.sub(r"\s+","",str(s)).lower() if s else ""
def scan_header(ws,row): return {c:(ws.cell(row=row,column=c).value or "") for c in range(1, ws.max_column+1)}

def excel_col_letters(idx):
    letters=""
    while idx:
        idx,rem = divmod(idx-1,26)
        letters=chr(65+rem)+letters
    return letters

def find_by_keywords(headers, keywords):
    if not keywords: return None
    keys=[normalize(k) for k in keywords]
    for c,text in headers.items():
        if any(k in normalize(text) for k in keys):
            return excel_col_letters(c)
    return None

def process_workbook(xlsx_path:Path,sheet:str,cfg:dict,overwrite=False):
    wb=load_workbook(str(xlsx_path))
    ws=wb[sheet] if sheet in wb.sheetnames else wb.active
    cols=dict(cfg.get("columns",{}))
    auto=cfg.get("autodetect",{})
    if auto.get("enabled",False):
        headers=scan_header(ws,int(auto.get("header_row",1)))
        for key in ["Y","Z","D","AD","AA","AB","AE","AF","AG"]:
            hit=find_by_keywords(headers,(auto.get("header_keywords") or {}).get(key,[]))
            if hit: cols[key]=hit
    start=int(cfg.get("start_row",2))
    stop=cfg.get("stop_row")
    max_row=ws.max_row if stop is None else min(ws.max_row,int(stop))
    fmt=cfg.get("percent_format","0.00%")
    for r in range(start,max_row+1):
        Y,Z,D,AD=[ws[f"{cols[k]}{r}"].value for k in ["Y","Z","D","AD"]]
        AA=safe_add(Y,Z); AB=safe_div(AA,D); AE=safe_add(AA,AD); AF=safe_sub(AE,D); AG=safe_div(AF,D)
        ws[f"{cols['AA']}{r}"].value=AA
        ws[f"{cols['AB']}{r}"].value=AB; ws[f"{cols['AB']}{r}"].number_format=fmt
        ws[f"{cols['AE']}{r}"].value=AE
        ws[f"{cols['AF']}{r}"].value=AF
        ws[f"{cols['AG']}{r}"].value=AG; ws[f"{cols['AG']}{r}"].number_format=fmt
    out=xlsx_path if overwrite else xlsx_path.with_name(xlsx_path.stem+"_processed.xlsx")
    wb.save(str(out)); return out

def main(argv=None):
    import sys
    p=argparse.ArgumentParser()
    p.add_argument("--input",required=True)
    p.add_argument("--sheet",required=True)
    p.add_argument("--config",required=True)
    p.add_argument("--overwrite",action="store_true")
    args=p.parse_args(argv)
    cfg=load_jsonc(Path(args.config))
    out=process_workbook(Path(args.input),args.sheet,cfg,overwrite=args.overwrite)
    print(f"Processed: {out}")
    return 0

if __name__=="__main__":
    raise SystemExit(main())
